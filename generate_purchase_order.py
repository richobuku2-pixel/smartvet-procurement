"""
SmartVet Africa — Purchase Order Generator
===========================================
Generates a formatted Excel purchase order for Global Vet (U) Ltd.

Usage (Claude Code / terminal):
    python generate_purchase_order.py
    python generate_purchase_order.py --order-no SV-GV-003 --output ./orders/
    python generate_purchase_order.py --quantities quantities.json

Dependencies:
    pip install openpyxl

Author:  SmartVet Africa dev team
Version: 1.0.0
"""

import argparse
import json
import os
import sys
from datetime import date
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.page import PageMargins
except ImportError:
    sys.exit("ERROR: openpyxl not installed. Run:  pip install openpyxl")


# ─── Colour palette ───────────────────────────────────────────────────────────
C = {
    "red_gv":    "8B0000",   # Global Vet dark red
    "gold_gv":   "C8973A",   # Global Vet gold
    "green_sv":  "2D5A2D",   # SmartVet green
    "dark":      "1A1A1A",
    "navy":      "1F3864",   # Vaccines
    "teal":      "1F5C6B",   # Respiratory
    "maroon":    "6B0F1A",   # Coccidiosis
    "olive":     "4A5E2A",   # Vitamins
    "slate":     "2F4F4F",   # Enteric
    "burnt":     "7B3F00",   # Dewormers
    "purple":    "4B2D7F",   # Disinfectants
    "star_hi":   "FF4444",   # ★★★
    "star_mid":  "FF9900",   # ★★
    "star_lo":   "888888",   # ★
}

# ─── Supplier & buyer details ─────────────────────────────────────────────────
SUPPLIER = {
    "name":    "Global Vet (U) Ltd",
    "address": "Plot 14/18 Nakivubo Place, Venus Plaza, Kampala",
    "tel":     "+256-41 4 507055  |  Mob: +256-77 6 421315",
    "email":   "globalvet@infocom.co.ug",
    "web":     "[www.globalvetug.com](https://www.globalvetug.com)",
}

BUYER = {
    "name":    "SmartVet Africa",
    "address": "Dark Store — Gulu, Northern Uganda",
    "contact": "Richard Obuku",
    "web":     "smartvetafrica.com",
}

# ─── Product catalogue ────────────────────────────────────────────────────────
# Structure: (section_header | product_dict)
# section_header: {"section": True, "title": str, "color": str}
# product_dict:   {
#     "id":          str,   -- unique product code
#     "name":        str,
#     "spec":        str,   -- pack size / specification
#     "indication":  str,   -- disease / use case
#     "priority":    str,   -- "★★★" | "★★" | "★"
#     "note":        str,   -- dosage / handling note
#     "color":       str,   -- inherits section colour
# }

CATALOGUE = [

    # ── SECTION 1: VACCINES ───────────────────────────────────────────────
    {"section": True, "color": C["navy"],
     "title": "SECTION 1: VACCINES FOR POULTRY — Routine Vaccination Programme"},

    {"id":"VAC-01","name":"BIO-VAC LA SOTA (Newcastle Disease Live Vaccine)",
     "spec":"1,000 doses/vial  |  Freeze-dried live",
     "indication":"Newcastle Disease (NCD)\n#1 killer of village poultry",
     "priority":"★★★","color":C["navy"],
     "note":"Via drinking water or eye-drop. Store -15°C to -4°C"},

    {"id":"VAC-02","name":"OL-VAC (Inactivated Newcastle Disease Vaccine)",
     "spec":"1,000 doses  |  Injectable inactivated",
     "indication":"Newcastle Disease — booster / inactivated option",
     "priority":"★★★","color":C["navy"],
     "note":"IM or SC injection. Useful where live vaccine insufficient"},

    {"id":"VAC-03","name":"BIO-VAC ND-IB (Newcastle + Infectious Bronchitis)",
     "spec":"1,000 doses/vial  |  Freeze-dried live combination",
     "indication":"Newcastle Disease + Infectious Bronchitis (IB/NCD)\nTwo diseases, one vaccination",
     "priority":"★★★","color":C["navy"],
     "note":"Via drinking water. Recommended for broilers week 2–3"},

    {"id":"VAC-04","name":"IBA-VAC ST (Gumboro — 1st Vaccination)",
     "spec":"1,000 doses/vial  |  Freeze-dried live",
     "indication":"Gumboro Disease (IBD)\nDevastates immune system in chicks",
     "priority":"★★★","color":C["navy"],
     "note":"Via drinking water. Administer day 10–14"},

    {"id":"VAC-05","name":"IBA-VAC (Gumboro — 2nd Vaccination)",
     "spec":"1,000 doses/vial  |  Freeze-dried live",
     "indication":"Gumboro Disease (IBD) — booster vaccination",
     "priority":"★★★","color":C["navy"],
     "note":"Via drinking water. Administer day 21–24"},

    {"id":"VAC-06","name":"BIO-MAREK HVT (Marek's Disease Vaccine)",
     "spec":"1,000 doses/vial  |  Include diluent",
     "indication":"Marek's Disease\nCauses tumours, paralysis, high mortality in layers",
     "priority":"★★★","color":C["navy"],
     "note":"SC injection — day-old chicks only. Full cold chain required"},

    {"id":"VAC-07","name":"BI-VAC 1° (Infectious Bronchitis — 1st Vaccination)",
     "spec":"1,000 doses/vial  |  Freeze-dried live",
     "indication":"Infectious Bronchitis (IB)\nRespiratory — reduces egg production",
     "priority":"★★","color":C["navy"],
     "note":"Via drinking water. Week 1"},

    {"id":"VAC-08","name":"BI-VAC 2° (Infectious Bronchitis — 2nd Vaccination)",
     "spec":"1,000 doses/vial  |  Freeze-dried live",
     "indication":"Infectious Bronchitis (IB) — booster",
     "priority":"★★","color":C["navy"],
     "note":"Via drinking water. Week 4–5"},

    {"id":"VAC-09","name":"VAIOL-VAC (Fowl Pox Vaccine)",
     "spec":"1,000 doses/vial  |  Freeze-dried live",
     "indication":"Fowl Pox\nCommon in tropical humid conditions",
     "priority":"★★","color":C["navy"],
     "note":"Wing-web stab method. Week 6–8"},

    {"id":"VAC-10","name":"SET-VAC (Fowl Typhoid Vaccine)",
     "spec":"500 doses or 1,000 doses  |  State size required",
     "indication":"Fowl Typhoid (Salmonella gallinarum)\nHigh mortality — common in Uganda",
     "priority":"★★","color":C["navy"],
     "note":"Drinking water or injection"},

    # ── SECTION 2: RESPIRATORY ────────────────────────────────────────────
    {"section": True, "color": C["teal"],
     "title": "SECTION 2: RESPIRATORY DISEASE TREATMENTS — CRD, Mycoplasmosis, IB"},

    {"id":"RES-01","name":"Erythrate 20% (Erythromycin thiocyanate 20%)",
     "spec":"100g sachet",
     "indication":"Chronic Respiratory Disease (CRD)\nMycoplasma gallisepticum — most common respiratory",
     "priority":"★★★","color":C["teal"],
     "note":"30g in 20L water for 5 days. First-line CRD treatment"},

    {"id":"RES-02","name":"Erythrate 35% (Erythromycin thiocyanate 35%)",
     "spec":"100g sachet",
     "indication":"CRD / Mycoplasmosis — higher concentration",
     "priority":"★★★","color":C["teal"],
     "note":"20g in 20L water for 5 days"},

    {"id":"RES-03","name":"Doxy Tylovet Plus (Doxycycline 200mg + Tylosin 100mg)",
     "spec":"1kg bag  |  Water soluble powder",
     "indication":"CRD, Mycoplasma, E.coli, Bordetella, Haemophilus, Salmonella\nBroad respiratory + GI combination",
     "priority":"★★★","color":C["teal"],
     "note":"1kg per 2,000–4,000L water for 3–5 days"},

    {"id":"RES-04","name":"Doxyvet 500 (Doxycycline Hyclate 500mg)",
     "spec":"100g sachet  |  Water soluble",
     "indication":"Mycoplasma gallisepticum / Respiratory infections",
     "priority":"★★","color":C["teal"],
     "note":"40g in 20L water for 1–7 days"},

    {"id":"RES-05","name":"Enrocure Solution (Enrofloxacin)",
     "spec":"100ml / 500ml / 1 Litre  |  State size",
     "indication":"Respiratory, enteric & urinary — broad-spectrum fluoroquinolone",
     "priority":"★★★","color":C["teal"],
     "note":"20ml in 20L water for 3–5 days"},

    {"id":"RES-06","name":"Enrotrill Solution (Enrofloxacin 200mg)",
     "spec":"Bottle  |  Higher concentration",
     "indication":"Respiratory, enteric & urinary — alternative Enrofloxacin",
     "priority":"★★","color":C["teal"],
     "note":"5ml in 20L water for 3–5 days"},

    {"id":"RES-07","name":"Gen Tylo (Gentamicin sulphate + Tylosin tartrate)",
     "spec":"Water soluble powder  |  100g sachet",
     "indication":"Wide-range respiratory, enteric & urinary bacterial infections",
     "priority":"★★","color":C["teal"],
     "note":"20g in 20L water for 5 consecutive days"},

    {"id":"RES-08","name":"Oxytetravet 25% (Oxytetracycline 25%)",
     "spec":"100g or 500g  |  State size",
     "indication":"Septicaemia, respiratory, urinary, enteric — gram +/−\nPullorum, Fowl Cholera, E.coli",
     "priority":"★★★","color":C["teal"],
     "note":"10g in 20L water for 5 days. High-volume workhorse antibiotic"},

    {"id":"RES-09","name":"Oxytetravet 50% (Oxytetracycline hydrochloride 50%)",
     "spec":"Water soluble powder",
     "indication":"Septicaemia, respiratory & GI infections in chicken",
     "priority":"★★","color":C["teal"],
     "note":"10g in 20L water for 5 days. Higher concentration"},

    {"id":"RES-10","name":"Oxy Plus (Oxytetracycline + Neomycin + Vit A, K)",
     "spec":"400g pack  |  Water soluble combination",
     "indication":"Respiratory + intestinal + soft tissue\nPullorum, Fowl Cholera, Fowl Typhoid, E.coli",
     "priority":"★★★","color":C["teal"],
     "note":"400g in 100L water for 5 days"},

    # ── SECTION 3: COCCIDIOSIS ────────────────────────────────────────────
    {"section": True, "color": C["maroon"],
     "title": "SECTION 3: COCCIDIOSIS — Most Common in Broilers Weeks 2–5"},

    {"id":"COC-01","name":"Amprocidia 60% (Amprolium Hydrochloride 60%)",
     "spec":"100g sachet  |  Water soluble",
     "indication":"Coccidiosis prevention & treatment\n#1 cause of diarrhoea + mortality in broilers",
     "priority":"★★★","color":C["maroon"],
     "note":"10g in 20L water for 5 consecutive days"},

    {"id":"COC-02","name":"Coccid Soluble Powder (Amprolium Hydrochloride 20%)",
     "spec":"Box of 10 sachets × 30g",
     "indication":"Coccidiosis prevention & treatment\nSachet-pack — ideal for smallholder farms",
     "priority":"★★★","color":C["maroon"],
     "note":"Severe: 1 tsp per 5L water for 5–7 days"},

    {"id":"COC-03","name":"Anticox (Diclazuril — Anticoccidial solution)",
     "spec":"Bottle  |  Liquid solution",
     "indication":"Coccidiosis — Diclazuril mode of action\nRotate with Amprolium to manage resistance",
     "priority":"★★","color":C["maroon"],
     "note":"20ml in 20L water for 3 days"},

    {"id":"COC-04","name":"Primovet (Sulphadiazine-Na + Trimethoprim)",
     "spec":"Water soluble powder",
     "indication":"Respiratory & GI diseases — E.coli, Pasteurella, Salmonella, Coccidiosis",
     "priority":"★★","color":C["maroon"],
     "note":"10g in 20L water for 5 consecutive days"},

    {"id":"COC-05","name":"S-Dime Solution (Sodium Sulfadimidine 16%)",
     "spec":"125ml or 1 Litre  |  State size",
     "indication":"Coccidiosis, Coryza, Fowl Cholera, Fowl Typhoid, Pullorum\nLiquid — easy accurate dosing",
     "priority":"★★★","color":C["maroon"],
     "note":"40ml in 5L water for 3–5 days"},

    # ── SECTION 4: VITAMINS & IMMUNITY ───────────────────────────────────
    {"section": True, "color": C["olive"],
     "title": "SECTION 4: VITAMINS, ELECTROLYTES & IMMUNE SUPPORT"},

    {"id":"VIT-01","name":"Multi-Aminovet",
     "spec":"Water soluble powder  |  100g sachet",
     "indication":"Vitamin deficiencies, growth, immunity, FCR, egg production\nAll stages — give before/after vaccinations",
     "priority":"★★★","color":C["olive"],
     "note":"10g (1 tbsp) in 20L water for 3–5 days"},

    {"id":"VIT-02","name":"Aminovet Solution",
     "spec":"1 Litre  |  Liquid",
     "indication":"Vitamin deficiencies, growth, immunity, FCR & egg production\nLiquid alternative",
     "priority":"★★★","color":C["olive"],
     "note":"5ml in 20L water"},

    {"id":"VIT-03","name":"Cosvita WS (Water Soluble Multivitamin)",
     "spec":"30g sachet",
     "indication":"Vitamin supplementation, retarded growth, convalescence\nCombine with antibiotics during treatment",
     "priority":"★★","color":C["olive"],
     "note":"1 sachet (30g) in 20L water for 5–7 days"},

    {"id":"VIT-04","name":"Poltricin Chick Formula",
     "spec":"Box of 5 × 25g sachets",
     "indication":"Chick starter — growth, CRD prevention, GI disease, stress\nSpecifically for growing chicks",
     "priority":"★★★","color":C["olive"],
     "note":"Starting chicks: 1 tbsp in 5L water for 7 days"},

    {"id":"VIT-05","name":"Poltricin Mayai Formula",
     "spec":"Box of 5 × 25g sachets",
     "indication":"Layer production — peak production, egg quality, FCR\nFor laying hens at point of lay",
     "priority":"★★","color":C["olive"],
     "note":"1 heaped tsp in 10L water at point of lay for 7 days"},

    # ── SECTION 5: ENTERIC / BULK ─────────────────────────────────────────
    {"section": True, "color": C["slate"],
     "title": "SECTION 5: ENTERIC & SYSTEMIC — BULK DARK STORE STOCK"},

    {"id":"ENT-01","name":"Oxytetravet 25% — Bulk (500g pack)",
     "spec":"500g pack  |  For dark store dispensing",
     "indication":"Septicaemia, E.coli, Salmonella, Pasteurella, Pullorum\nHigh-frequency — order in bulk",
     "priority":"★★★","color":C["slate"],
     "note":"10g in 20L water for 5 days. Dispense in 100g sachets"},

    {"id":"ENT-02","name":"Enrocure Solution — Bulk (1 Litre)",
     "spec":"1 Litre  |  For dark store dispensing into 100ml bottles",
     "indication":"Respiratory, enteric & urinary — most versatile single antibiotic\nDispense to paravets in 100ml units",
     "priority":"★★★","color":C["slate"],
     "note":"20ml in 20L water. Repackage in 100ml bottles for paravets"},

    # ── SECTION 6: DEWORMERS ──────────────────────────────────────────────
    {"section": True, "color": C["burnt"],
     "title": "SECTION 6: DEWORMERS — Poultry"},

    {"id":"DEW-01","name":"Ascarex D Worm Powder (Piperazine Citrate)",
     "spec":"Box of 10 sachets × 30g  |  Water soluble",
     "indication":"Roundworms (Ascaridiasis) in poultry\nHighly prevalent in free-range & semi-intensive flocks",
     "priority":"★★★","color":C["burnt"],
     "note":"30g in 20–30L water per 100 birds. 1-day treatment only"},

    {"id":"DEW-02","name":"Piperamentic (Piperazine as citrate)",
     "spec":"100g or 500g  |  State size required",
     "indication":"Roundworms — Ascaridiasis & other roundworms in poultry\nBulk pack for dark store dispensing",
     "priority":"★★★","color":C["burnt"],
     "note":"40g in 20L water for 1 day only. Do not exceed"},

    # ── SECTION 7: DISINFECTANTS ──────────────────────────────────────────
    {"section": True, "color": C["purple"],
     "title": "SECTION 7: DISINFECTANTS, WOUND CARE & BIOSECURITY"},

    {"id":"DIS-01","name":"Oytetravet Aerosol (Oxytetracycline + purple dye spray)",
     "spec":"Aerosol can",
     "indication":"Wound treatment & fly repellent\nDebeaking wounds, cannibalism injuries, scaly leg",
     "priority":"★★","color":C["purple"],
     "note":"Spray directly on wound. Purple dye helps monitor healing"},

    {"id":"DIS-02","name":"Disinfectant — advise equivalent to Norocleanse",
     "spec":"1 Litre / 5 Litre  |  State available brands",
     "indication":"Poultry house disinfection, water system flushing, equipment\nBiosecurity between flocks",
     "priority":"★★★","color":C["purple"],
     "note":"Please advise preferred product, dilution rate & cost"},

    # ── SECTION 8: EQUIPMENT ──────────────────────────────────────────────
    {"section": True, "color": C["dark"],
     "title": "SECTION 8: EQUIPMENT & ADMINISTRATION SUPPLIES"},

    {"id":"EQP-01","name":"HSW ECO-MATIC Automatic Syringe",
     "spec":"0.3ml / 1ml / 2ml / 5ml  |  State size",
     "indication":"Vaccine & drug administration — injectable products\nEssential for Marek's and OL-VAC",
     "priority":"★★★","color":C["dark"],
     "note":"State sizes required. One per paravet field kit"},

    {"id":"EQP-02","name":"HSW ROUX-REVOLVER Automatic Syringe",
     "spec":"10ml / 30ml / 50ml  |  State size",
     "indication":"Large-volume drug administration",
     "priority":"★★","color":C["dark"],
     "note":"For bulk antibiotic injection in larger farms"},

    {"id":"EQP-03","name":"EUROPLEX Automatic Drencher",
     "spec":"30ml",
     "indication":"Oral deworming / drenching — poultry & small animals",
     "priority":"★★","color":C["dark"],
     "note":"Accurate dosing for deworming rounds"},

    {"id":"EQP-04","name":"Leur Lock Hard Plastic Syringes",
     "spec":"Assorted sizes — state volumes required",
     "indication":"General vaccine and drug administration",
     "priority":"★★","color":C["dark"],
     "note":"Order adequate supply per visit cycle"},
]


# ─── Style helpers ────────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def _border(style="all", color="CCCCCC"):
    s = Side(style="thin", color=color)
    if style == "all":
        return Border(left=s, right=s, top=s, bottom=s)
    if style == "bot":
        return Border(bottom=s)
    return Border()

def _align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _write(ws, cell_or_range, value, **styles):
    """Write value to a cell or merged range with optional styles."""
    if ":" in str(cell_or_range):
        ws.merge_cells(cell_or_range)
        cell = ws[cell_or_range.split(":")[0]]
    else:
        cell = ws[cell_or_range]
    cell.value = value
    if "font"   in styles: cell.font      = styles["font"]
    if "fill"   in styles: cell.fill      = styles["fill"]
    if "align"  in styles: cell.alignment = styles["align"]
    if "border" in styles: cell.border    = styles["border"]
    if "num"    in styles: cell.number_format = styles["num"]
    return cell


# ─── Section builders ─────────────────────────────────────────────────────────

def build_header(ws, order_no, order_date, payment_terms="30 days net"):
    """Write the two-party header block (rows 1–8)."""
    row_heights = {1: 8, 2: 34, 3: 8, 4: 17, 5: 17, 6: 17, 7: 17, 8: 8}
    for r, h in row_heights.items():
        ws.row_dimensions[r].height = h

    # Title bar
    _write(ws, "A1:I1", "", fill=_fill(C["red_gv"]))
    _write(ws, "A2:E2", "PURCHASE ORDER",
           font=_font(True, "FFFFFF", 18),
           fill=_fill(C["red_gv"]),
           align=_align("left", "center", False))
    _write(ws, "F2:I2", "Global Vet (U) Ltd  ·  Veterinary Pharmaceuticals",
           font=_font(True, C["gold_gv"], 11),
           fill=_fill(C["red_gv"]),
           align=_align("right", "center", False))
    _write(ws, "A3:I3", "", fill=_fill(C["gold_gv"]))

    # Info rows
    info = [
        (4, f"SUPPLIER:  {SUPPLIER['name']}  |  {SUPPLIER['address']}  |  Tel: {SUPPLIER['tel']}",
            "ORDER NO:", order_no),
        (5, f"Email: {SUPPLIER['email']}  |  {SUPPLIER['web']}",
            "ORDER DATE:", order_date),
        (6, f"SHIP TO:  {BUYER['name']}  |  {BUYER['address']}",
            "REQUIRED BY:", "To be confirmed"),
        (7, f"Contact:  {BUYER['contact']}  |  {BUYER['name']}",
            "PAYMENT TERMS:", payment_terms),
    ]
    for r, left_val, key, val in info:
        _write(ws, f"A{r}:E{r}", left_val, font=_font(False, C["dark"], 9), align=_align())
        _write(ws, f"F{r}:G{r}", key,      font=_font(True, C["dark"], 9),  align=_align("right"))
        _write(ws, f"H{r}:I{r}", val,      font=_font(True, "0000FF", 9),   align=_align("left"))

    _write(ws, "A8:I8", "", fill=_fill("DDDDDD"))


def build_legend(ws, row=9):
    """Write the priority legend row."""
    ws.row_dimensions[row].height = 14
    _write(ws, f"A{row}:I{row}",
           "  ★★★ = High demand / critical    ★★ = Important    ★ = Useful    "
           "Yellow = SmartVet fills Qty    Red tint = Global Vet fills Unit Price    "
           "Amount = auto-calculated",
           font=_font(False, "FFFFFF", 8),
           fill=_fill("333333"),
           align=_align("left"))


def build_table_header(ws, row=10):
    """Write the column header row."""
    ws.row_dimensions[row].height = 26
    headers = [
        ("A", "#"), ("B", "Product Name"), ("C", "Pack Size / Specification"),
        ("D", "Disease / Indication"), ("E", "Priority"), ("F", "Qty"),
        ("G", "Unit Price\n(UGX)"), ("H", "Amount\n(UGX)"), ("I", "Notes / Dosage"),
    ]
    for col, label in headers:
        c = ws[f"{col}{row}"]
        c.value = label
        c.font = _font(True, "FFFFFF", 10)
        c.fill = _fill(C["dark"])
        c.alignment = _align("center")
        c.border = _border()


def build_products(ws, catalogue, quantities=None, start_row=11):
    """
    Write all product rows and section headers.

    quantities: dict mapping product id → qty  e.g. {"VAC-01": 10, "RES-01": 5}
                If None, all Qty cells are left blank (yellow) for manual entry.

    Returns the list of data row numbers (for SUM formula).
    """
    row = start_row
    item_rows = []
    quantities = quantities or {}

    for entry in catalogue:
        if entry.get("section"):
            # ── Section header row ─────────────────────────────────────────
            ws.row_dimensions[row].height = 18
            ws.merge_cells(f"A{row}:I{row}")
            c = ws[f"A{row}"]
            c.value = f"  {entry['title']}"
            c.font = _font(True, "FFFFFF", 10)
            c.fill = _fill(entry["color"])
            c.alignment = _align("left")
            c.border = _border()
        else:
            # ── Product row ────────────────────────────────────────────────
            ws.row_dimensions[row].height = 36
            item_rows.append(row)
            shaded = len(item_rows) % 2 == 0
            row_bg = "F8F5F0" if shaded else "FFFFFF"

            star = entry["priority"]
            star_color = (C["star_hi"] if "★★★" in star
                          else C["star_mid"] if "★★" in star
                          else C["star_lo"])

            qty_val = quantities.get(entry["id"], "")

            cell_data = {
                "A": (entry.get("id", ""),   _font(False, "999999", 8),     _align("center"), row_bg,    None),
                "B": (entry["name"],          _font(True,  C["dark"], 9),    _align("left"),   row_bg,    None),
                "C": (entry["spec"],          _font(False, "555555", 8),     _align("left"),   row_bg,    None),
                "D": (entry["indication"],    _font(False, "1A1A6B", 8),     _align("left"),   row_bg,    None),
                "E": (star,                   _font(True,  star_color, 9),   _align("center"), row_bg,    None),
                "F": (qty_val,                _font(True,  "0000FF", 11),    _align("center"), "FFFDE7",  "#,##0"),
                "G": ("",                     _font(True,  "AA0000", 10),    _align("right"),  "FFF0F0",  "#,##0"),
                "H": (f"=IF(AND(ISNUMBER(F{row}),ISNUMBER(G{row})),F{row}*G{row},\"\")",
                                              _font(True,  C["dark"], 10),   _align("right"),  row_bg,    "#,##0"),
                "I": (entry["note"],          _font(False, "666666", 8),     _align("left"),   row_bg,    None),
            }

            for col, (val, font, align, bg, num_fmt) in cell_data.items():
                c = ws[f"{col}{row}"]
                c.value = val
                c.font = font
                c.alignment = align
                c.fill = _fill(bg)
                c.border = _border()
                if num_fmt:
                    c.number_format = num_fmt

        row += 1

    return item_rows


def build_totals(ws, item_rows, row):
    """Write subtotal, VAT, and grand total rows."""
    ws.row_dimensions[row].height = 6
    row += 1

    sub_row = row
    f_sum = f"=SUM(H{item_rows[0]}:H{item_rows[-1]})" if item_rows else "=0"

    def _tot_row(r, label, formula, bg, label_color, val_color, h=22, bold_val=False):
        ws.row_dimensions[r].height = h
        ws.merge_cells(f"A{r}:G{r}")
        c = ws[f"A{r}"]
        c.value = label
        c.font = _font(True, label_color, 10)
        c.fill = _fill(bg)
        c.alignment = _align("right")
        c.border = _border()
        v = ws[f"H{r}"]
        v.value = formula
        v.font = _font(bold_val, val_color, 11 if bold_val else 10)
        v.fill = _fill(bg)
        v.alignment = _align("right")
        v.number_format = "#,##0"
        v.border = _border()
        ws[f"I{r}"].fill = _fill(bg)
        ws[f"I{r}"].border = _border()

    _tot_row(row, "SUBTOTAL (UGX)", f_sum, "E8F0E8", C["green_sv"], C["green_sv"])
    row += 1

    vat_row = row
    _tot_row(row, "VAT / TAX — Global Vet to confirm applicable rate", "", "FFF0F0", "AA0000", "AA0000")
    ws[f"H{row}"].value = ""
    ws[f"H{row}"].fill = _fill("FFF0F0")
    ws[f"H{row}"].font = _font(True, "AA0000", 10)
    row += 1

    _tot_row(row,
             "GRAND TOTAL (UGX)",
             f"=H{sub_row}+IF(ISNUMBER(H{vat_row}),H{vat_row},0)",
             C["dark"], "FFFFFF", C["gold_gv"], h=28, bold_val=True)
    row += 1

    return row


def build_footer(ws, row, order_date):
    """Write completion instructions and document footer."""
    ws.row_dimensions[row].height = 6
    row += 1

    ws.row_dimensions[row].height = 16
    ws.merge_cells(f"A{row}:I{row}")
    c = ws[f"A{row}"]
    c.value = ("  COMPLETION:  Yellow (Qty) → SmartVet fills before sending  |  "
               "Red tint (Unit Price) → Global Vet fills and returns  |  "
               "Amount column auto-calculates")
    c.font = _font(False, "FFFFFF", 8)
    c.fill = _fill("333333")
    c.alignment = _align("left")
    row += 1

    ws.row_dimensions[row].height = 16
    ws.merge_cells(f"A{row}:I{row}")
    c = ws[f"A{row}"]
    c.value = ("  NOTES TO GLOBAL VET:  Please confirm stock availability, note any "
               "substitutions or out-of-stock items, and advise on cold chain handling "
               "for vaccines on delivery to Gulu.")
    c.font = _font(False, "444444", 8)
    c.fill = _fill("F5F5F5")
    c.alignment = _align("left", wrap=True)
    row += 1

    ws.row_dimensions[row].height = 14
    ws.merge_cells(f"A{row}:I{row}")
    c = ws[f"A{row}"]
    c.value = (f"  {BUYER['name']}  ·  Prepared {order_date}  ·  "
               f"{BUYER['address']}  ·  {BUYER['web']}")
    c.font = _font(False, "999999", 8)
    c.alignment = _align("center")


def set_column_widths(ws):
    widths = {"A": 7, "B": 28, "C": 26, "D": 22,
              "E": 9, "F": 9, "G": 16, "H": 16, "I": 18}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def configure_print(ws):
    ws.freeze_panes = "A11"
    ws.print_title_rows = "1:10"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.6, bottom=0.6)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9  # A4


# ─── Main generator ───────────────────────────────────────────────────────────

def generate_order(
    output_path: str = "SmartVet_Purchase_Order.xlsx",
    order_no: str = "SV-GV-001",
    order_date: str = None,
    payment_terms: str = "30 days net",
    quantities: dict = None,
) -> str:
    """
    Generate the SmartVet purchase order Excel file.

    Args:
        output_path:   Full path (including filename) for the output file.
        order_no:      Purchase order reference number.
        order_date:    Date string (e.g. "23 May 2026"). Defaults to today.
        payment_terms: Payment terms string.
        quantities:    Dict mapping product ID → integer quantity.
                       e.g. {"VAC-01": 10, "RES-01": 5}
                       Products not in dict will have blank Qty cells.

    Returns:
        Absolute path of the saved file.
    """
    order_date = order_date or date.today().strftime("%d %B %Y")

    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Order"

    set_column_widths(ws)
    build_header(ws, order_no, order_date, payment_terms)
    build_legend(ws, row=9)
    build_table_header(ws, row=10)
    item_rows = build_products(ws, CATALOGUE, quantities=quantities, start_row=11)
    next_row  = 11 + len(CATALOGUE)
    next_row  = build_totals(ws, item_rows, next_row)
    build_footer(ws, next_row, order_date)
    configure_print(ws)

    # Ensure output directory exists
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return str(Path(output_path).resolve())


# ─── CLI entry point ──────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Generate a SmartVet Africa purchase order for Global Vet (U) Ltd.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python generate_purchase_order.py
  python generate_purchase_order.py --order-no SV-GV-003 --output ./orders/
  python generate_purchase_order.py --quantities quantities.json

quantities.json format:
  {
    "VAC-01": 10,
    "VAC-04": 20,
    "RES-01": 15,
    "VIT-01": 30
  }

Product IDs follow the pattern: VAC-01..10, RES-01..10, COC-01..05,
VIT-01..05, ENT-01..02, DEW-01..02, DIS-01..02, EQP-01..04
        """,
    )
    parser.add_argument("--order-no",  default="SV-GV-001",
                        help="Purchase order reference number (default: SV-GV-001)")
    parser.add_argument("--output",    default=".",
                        help="Output directory or full file path (default: current directory)")
    parser.add_argument("--date",      default=None,
                        help="Order date string e.g. '23 May 2026' (default: today)")
    parser.add_argument("--payment",   default="30 days net",
                        help="Payment terms (default: '30 days net')")
    parser.add_argument("--quantities",default=None,
                        help="Path to JSON file mapping product IDs to quantities")

    args = parser.parse_args()

    # Resolve output path
    out = Path(args.output)
    if out.suffix.lower() == ".xlsx":
        output_path = str(out)
    else:
        fname = f"SmartVet_PO_{args.order_no}_{date.today().strftime('%Y%m%d')}.xlsx"
        output_path = str(out / fname)

    # Load quantities if supplied
    quantities = None
    if args.quantities:
        try:
            with open(args.quantities) as f:
                quantities = json.load(f)
            print(f"Loaded quantities from {args.quantities}")
        except (FileNotFoundError, json.JSONDecodeError) as e:
            sys.exit(f"ERROR loading quantities file: {e}")

    print(f"Generating purchase order {args.order_no} ...")
    saved = generate_order(
        output_path=output_path,
        order_no=args.order_no,
        order_date=args.date,
        payment_terms=args.payment,
        quantities=quantities,
    )
    print(f"✓  Saved:  {saved}")


if __name__ == "__main__":
    main()